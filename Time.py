#!usr/bin/python

import argparse
import datetime
import logging
import openpyxl
import os
import re
import subprocess

from jinja2 import Environment, FileSystemLoader
from openpyxl.styles import (Alignment, Color, fills, PatternFill)


class Course:

    def __init__(self):
        self.abbrev = None
        self.name = None
        self.tutor = None
        self.location = None
        self.count = 0
        self.count_to_method = {0:self.set_abbrev, 1:self.set_name,
                              2:self.set_tutor, 3:self.set_location}

    def record(self, entry):
        self.count_to_method[self.count](entry)

    def set_abbrev(self, abbrev):
        self.abbrev = abbrev
        self.count += 1

    def set_name(self, name):
        self.name = name
        self.count += 1

    def set_tutor(self, tutor):
        self.tutor = tutor
        self.count += 1

    def set_location(self, loc):
        self.location = loc
        self.count += 1

class Lecture:

    def __init__(self, course, date, time, year):
        self.abbrev = course
        self.date = date
        self.time = time
        self.year = year

    def set_location(self, loc):
        self.location = loc

    def set_name(self, name):
        self.name = name

    def __repr__(self):
        properties = [self.abbrev, self.date, self.time, self.year]
        return 'Lecture({0}, {1}, {2}, Year {3})'.format(*properties)

def fuck_timetable(year_sheet, args):

    """Removes irrelevant courses, changes time format and parses top RH
    corner description. 
    
    Assumes info box in top RH corner is bounded on L & R side by cells
    containing only the word CUT and END_CUT, respectively."""

    course_list = dict()
    in_cut = False

    for row in year_sheet:
        for cell in row:

            if not isinstance(cell.value, basestring):
                if cell.value is not None:
                    derryck.debug('NOT BASESTRING: {0}'.format(cell.value))
                continue

            cell_contains = {val.strip(' ') for val in cell.value.split()}
            # Parse RH corner box for course names, abbrevs and locations
            if 'CUT' in cell_contains:
                in_cut = True
                course = Course()
                continue
            if 'END_CUT' in cell_contains:
                in_cut = False
                course_list[course.abbrev] = course
                continue
            if in_cut:
                derryck.info('Recording {0}'.format(cell.value))
                course.record(cell.value)
                continue
            # Ignore column titles
            if any(['Week' in cell_contains, 'Time' in cell_contains,
                    'Mon' in cell_contains, 'Tues' in cell_contains,
                    'Wed' in cell_contains, 'Thur' in cell_contains,
                    'Fri' in cell_contains]):
                continue
            # Adjust formatting for time
            if re.search('[0-9]+\-[0-9]+', cell.value):
                start_time = cell.value.strip(' ="').split('-')[0]
                if start_time.startswith('0'):
                    start_time = start_time[1:]
                # Check that it is a pure numeric value
                elif start_time[0].isalpha():
                    continue
                # 24-hour format
                elif int(start_time) < 9:
                    start_time = str(int(start_time)+12)
                start_time += ':00'
                cell.value = start_time
                continue

            if not isinstance(args.courses, list):
                courses = set([unicode(args.courses)])
            else:
                courses = set(unicode(course) for course in args.courses)
            courses_in_cell = {item for item in cell_contains if 
                               item.strip('()[]') in courses}
            if len(courses_in_cell) > 0:
                cell.value = ' '.join(courses_in_cell)
                cell.alignment = Alignment(horizontal='center')
                continue
            cell.value = None
            cell.fill = PatternFill(patternType=fills.FILL_SOLID,
                                    fgColor=Color('FFFFFF'))

    timetable.save(os.path.abspath(args.path))
    derryck.info(course_list)
    return course_list

def move_to_iCal(year_sheet, template, course_list, args):

    """Runs through sheet again and creates iCal events if cell value
    encountered."""

    in_cut = False

    for row in year_sheet:
        for cell in row:
            
            if cell.value is None:
                continue
            
            try:
                cell_contains = {val.strip(' ') for val in cell.value.split()}
            # Some entries will not have a split method. Don't want to convert
            # these to strings
            except AttributeError:
                cell_contains = []
            
            if 'CUT' in cell_contains:
                in_cut = True
                cell.value = None
                continue
            if 'END_CUT' in cell_contains:
                in_cut = False
                cell.value = None
                continue
            if in_cut:
                continue
            # Ignore column titles
            if any(['Week' in cell_contains, 'Time' in cell_contains,
                    'Mon' in cell_contains, 'Tues' in cell_contains,
                    'Wed' in cell_contains, 'Thur' in cell_contains,
                    'Fri' in cell_contains]):
                continue

            # Obtain date and lecture start time
            if isinstance(cell.value, datetime.datetime):
                lec_date = cell.value
                continue
            elif re.search('[0-9]+\:00', cell.value):
                start_time = cell.value.strip(' ="') + ':00'
                lec_time = cell.value 
            
            # Look for lectures
            courses_in_cell = {item for item in cell_contains if
                               item.strip('()[]') in args.courses}
            # There should only be one item in that set but written
            # here for the case of multiple items
            for lecture in courses_in_cell:
                lecture = Lecture(lecture, lec_date, lec_time, args.year)
                lecture.set_name(course_list[lecture.abbrev.strip('[]')].name)
                lecture.set_location(course_list[lecture.abbrev.strip('[]')].location)
                derryck.info('Creating lecture {0}'.format(lecture))
                # Populate applescript template/pass these values to
                # applescript
                with open('tmpCalEventScript.scpt', 'w') as fh:
                    fh.write(template.render(lecture.__dict__))
                    subprocess.call(['osascript', fh.name])

if __name__ == '__main__':

    # Argument parsing
    derryck = logging.getLogger('Derryck')
    derryck.addHandler(logging.StreamHandler())
    
    desc_str = """Removes all but specified courses from Derryck-issued
                  timetables, doing so on a row-by-row basis. Cells in a given
                  row bounded by CUT and END_CUT will be ignored. Requires
                  openpyxl."""
    ep_str = 'Example: "python <path_to_timetable> 4 <path_to_template> -c GR QO" removes all but GR, QO.'
    parser = argparse.ArgumentParser(description=desc_str, epilog=ep_str)
    parser.add_argument('path', help='Path to timetable.')
    parser.add_argument('year', type=int, help='Year of study.')
    parser.add_argument('template', help='Path to applescript template.')
    parser.add_argument('--courses', '-c', metavar='C', nargs='+',
                        help='Relevant courses (abbrevs).')
    parser.add_argument('--verbose', '-v', action='store_true')
    args = parser.parse_args()
    
    log_level = logging.WARN
    if args.verbose:
        log_level = logging.DEBUG
    derryck.setLevel(log_level)

    # Obtain relevant Excel sheet
    sheet_number = args.year
    if sheet_number == 4:
        sheet_number = 3

    timetable = openpyxl.load_workbook(os.path.abspath(args.path))
    year_sheet = timetable[timetable.get_sheet_names()[sheet_number-1]]

    # Obtain applescript template
    env = Environment(loader=FileSystemLoader('.'))
    template = env.get_template(args.template)
    derryck.info('Loading applescript template {0}.'.format(template.filename))

    course_list = fuck_timetable(year_sheet, args)
    move_to_iCal(year_sheet, template, course_list, args)
